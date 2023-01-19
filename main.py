from nis import match
import openpyxl
import datetime
from datetime import date, datetime, timedelta
from openpyxl.styles import PatternFill
import re
import configparser

def check_dates(file_path, dates_sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.get_sheet_by_name(dates_sheet_name)
    columns = ['D', 'E', 'F']
    today = datetime.now().date()
    week_from_now = today + timedelta(days=30)
    matching_cells = []
    late_squad_cells = []

    for column in columns:
        for row in range(4, sheet.max_row + 1):
            cell = sheet[column + str(row)]
            if cell.value and isinstance(cell.value, datetime) and cell.value.date() <= week_from_now:
                matching_cells.append(cell)

    for match in matching_cells:
        late_cell = sheet['B' + str(match.row)]
        late_squad_cells.append(late_cell)

    return late_squad_cells, matching_cells



def check_team(file_path,team_sheet ,late_squad_cells):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Get the sheet2
    sheet2 = wb[team_sheet]
    # Initialize a dictionary to store the cells with the same content as the key
    teams = {}
    # Iterate over the late squad cells
    for late_squad_cell in late_squad_cells:
        # Get the corresponding cell in column I on sheet2
        condition_cell = sheet2['J' + str(late_squad_cell.row)]
        # Get the value of the condition cell
        condition = condition_cell.value
        if condition:
            team_cell = sheet2['G' + str(late_squad_cell.row)]
            # Get the value of the team cell
            team = team_cell.value
            # Check if the team is already in the dictionary
            if late_squad_cell.value in teams:
                # If it is, append the cell to the corresponding list
                teams[late_squad_cell.value].append(team)
            else:
                # If it's not, create a new key in the dictionary with the team as the key and the late squad cell as the value
                teams[late_squad_cell.value] = [team]
        if str(condition) == 'None':
            if late_squad_cell.value in teams:
                # If it is, append the cell to the corresponding list
                teams[late_squad_cell.value].append('None')
            else:
                # If it's not, create a new key in the dictionary with the team as the key and the late squad cell as the value
                teams[late_squad_cell.value] = ['None']
    return teams


    
def compare_teams(matching_cells, late_squad_cells):
    teams_late_dates = {}
    matching_cells_corrected = []
    late_squad_corrected= []
    
    for coordinate in matching_cells:
        #print(coordinate.coordinate)
        if 'D' in coordinate.coordinate:
            sprint_type = 'sprint1'
        if 'E' in coordinate.coordinate:
            sprint = ['lol']
            sprint_type = 'sprint2'
        if 'F' in coordinate.coordinate:
            sprint = ['test']
            sprint_type = 'sprint3'
        matching_cells_corrected.append((coordinate.value , sprint_type))

        for late_squad in late_squad_cells:
            late_squad_corrected.append(late_squad.value)

    teams_late_dates = {key: value for key, value in zip(late_squad_corrected, matching_cells_corrected)}
    return teams_late_dates

def find_cells(teams_late_dates, file_path, tracker_sheet):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[tracker_sheet]
    date_row = sheet[2]
    team_and_date = {}
    testList = []

    #find the correct team cell in the sheet
    for team in teams_late_dates.keys():
        for cell_c in sheet['C']:
            if cell_c.value:
                testList.append(cell_c.value)

            if cell_c.value == team:

                for cell in date_row:
                        if cell.value == teams_late_dates[team][0]:
                            if cell_c.coordinate in team_and_date:
                                team_and_date[cell_c.coordinate].append(cell.coordinate)
                                team_and_date[cell_c.coordinate].append(teams_late_dates[team][1])

                            else:
                                team_and_date[cell_c.coordinate] = [cell.coordinate]
                                team_and_date[cell_c.coordinate].append(teams_late_dates[team][1])

    return team_and_date


def find_team_role(team_and_date,file_path, team_sheet, tracker_sheet):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(team_sheet)
    tracker_sheet = wb.get_sheet_by_name(tracker_sheet)
    team_roles = {}

    for key in team_and_date.keys():
        key = tracker_sheet[key].value
        for row in range(1, sheet.max_row + 1):

            if key == sheet.cell(row=row, column=4).value:

                if str((sheet.cell(row=row, column=10).value)) != 'None':

                    if key in team_roles:
                        team_roles[key].append((sheet.cell(row=row, column=7).value))
                        
                    else:
                        team_roles[key] = [key]
                        team_roles[key].append((sheet.cell(row=row, column=7).value))
                
                else:
                    pass

    team_roles = {k: v[1:] for k, v in team_roles.items()}
    return team_roles



def sort_team(file_path, tracker_sheet, team_and_date, team_roles):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(tracker_sheet)
    sprint4 = []

    for team in team_and_date:
        team = sheet[team]
        for team_name in team_roles:
            if team_name == team.value:

                team_and_date[team.coordinate].append(team_roles[team_name])

    for squad_cell_to_fill in team_and_date:
    
        if team_and_date[squad_cell_to_fill][1] == 'sprint1':
            team_and_date[squad_cell_to_fill].append(['Designer', 'Senior Dev', 'Dev', 'Test'])

        if team_and_date[squad_cell_to_fill][1] == 'sprint2':
            team_and_date[squad_cell_to_fill].append(['Dev', 'Dev', 'Dev','Test', 'Test', 'Senior Dev', 'Designer', 'Designer' ])

        if team_and_date[squad_cell_to_fill][1] == 'sprint3':
            team_and_date[squad_cell_to_fill].append(['Designer', 'Designer', 'Senior Dev', 'Dev', 'Dev', 'Dev', 'Test', 'Test', 'Scrum Master'])


        if team_and_date[squad_cell_to_fill][1] == 'sprint4':
            team_and_date[squad_cell_to_fill].append(sprint4)


    return team_and_date
            

def colour_now(file_path, tracker_sheet, team_date_role):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(tracker_sheet)


    for squad_cell in team_date_role:
        
        start_row = sheet[squad_cell].row
        end_row = start_row + 20
        for default_row in range(start_row,end_row+1):
            cell_value = sheet.cell(row=default_row,column=7)
            red_fill = PatternFill(start_color="FF0000" , fill_type="solid")
            green_fill = PatternFill(start_color="0000FF00" , fill_type="solid")
            cell_string = str(cell_value.coordinate)
            row = re.findall('\d+', cell_string)
            row =  row[0]
            row2 = str(int(row) +1)
            column = " ".join(re.findall("[a-zA-Z]+", team_date_role[squad_cell][0] )) 
            
            if sheet[column + row].fill.start_color.index == '00000000':

                if str(cell_value.value) in team_date_role[squad_cell][3] and str(cell_value.value) in team_date_role[squad_cell][2]:
                    sheet[column + row].fill = green_fill
                    print(squad_cell)
                    team_date_role[squad_cell][3].remove(str(cell_value.value))
                    team_date_role[squad_cell][2].remove(str(cell_value.value))
                    
                if str(cell_value.value) in team_date_role[squad_cell][3] and str(cell_value.value) in team_date_role[squad_cell][2]:
                    if sheet[column + row].fill.start_color.index == '0000FF00':
                        sheet[column + row2].fill = green_fill
                        team_date_role[squad_cell][2].remove(str(cell_value.value))
                        team_date_role[squad_cell][3].remove(str(cell_value.value))


    for squad_cell in team_date_role:
        start_row = sheet[squad_cell].row
        end_row = start_row + 20
        for default_row in range(start_row,end_row+1):
                
                if sheet['G' + str(default_row)].value in team_date_role[squad_cell][3]:

                    cell = sheet.cell(row=default_row,column=7)
                    cell = str(cell.coordinate)
                    row = re.findall('\d+', cell_string)
                    default_row =  str(default_row)
                    column = " ".join(re.findall("[a-zA-Z]+", team_date_role[squad_cell][0] )) 

                    if sheet[column + default_row].fill.start_color.index == '00000000':
                        sheet[column + default_row].fill = red_fill
                        team_date_role[squad_cell][3].remove(sheet['G' + str(default_row)].value)
                            
        
    wb.save(config['User_Settings']['updated_sheet_name'])


if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read('config.ini')
    file_path = config['User_Settings']['file_path']
    dates_sheet_name = config['Dev_Settings']['dev_sheet']
    team_sheet = config['User_Settings']['teams_sheet']
    tracker_sheet = config['User_Settings']['Squad_tracking_sheet']

    late_squad_cells, matching_cells = check_dates(file_path, dates_sheet_name)

    if late_squad_cells:
        late_squads = check_team(file_path, team_sheet,late_squad_cells)
        teams_late_dates = compare_teams(matching_cells, late_squad_cells)
        team_and_date = find_cells(teams_late_dates, file_path, tracker_sheet)
        team_roles = find_team_role(team_and_date, file_path, team_sheet, tracker_sheet)
        team_date_role = sort_team(file_path, tracker_sheet, team_and_date, team_roles)
        colour_now(file_path, tracker_sheet, team_date_role)

