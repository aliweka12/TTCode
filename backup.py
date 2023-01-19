from nis import match
import openpyxl
import datetime
from datetime import date, datetime, timedelta
from openpyxl.styles import PatternFill
import re




def check_dates(file_path, dates_sheet_name):
    # Open the workbook and return the sheet
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.get_sheet_by_name(dates_sheet_name)
    # Define the columns we want to check
    columns = ['D', 'E', 'F']
    # Get the current date
    today = datetime.now().date()
    # Get the date a week from now
    week_from_now = today + timedelta(days=30)
    # Initialize a list to store the cells that meet the criteria
    matching_cells = []
    # Iterate over the columns
    for column in columns:
        # Iterate over the rows starting from 4
        for row in range(4, sheet.max_row + 1):
            # Get the cell in the current column and row
            cell = sheet[column + str(row)]
            # Check if the cell's value is a date and if it is a week or less from now
            if cell.value and isinstance(cell.value, datetime) and cell.value.date() <= week_from_now:
                #print(cell.coordinate)

                # If it is, add it to the list of matching cells
                matching_cells.append(cell)
    # Initialize a list to store the corresponding cells from column B
    late_squad_cells = []
    # Iterate over the matching cells
   # for value in matching_cells:
    #    print(value.coordinate)
    for match in matching_cells:
        # Get the cell in column B at the same row as the matching cell
        late_cell = sheet['B' + str(match.row)]
        # Add the cell to the list
        late_squad_cells.append(late_cell)
    # Return the list of corresponding cells from column B


    return late_squad_cells, matching_cells



def check_team(file_path,team_sheet ,late_squad_cells):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Get the sheet2
    sheet2 = wb[team_sheet]
    # Initialize a dictionary to store the cells with the same content as the key
    teams = {}
    # Iterate over the late squad cells
    for late_squad_cell in late_squad_cells:
        # Get the corresponding cell in column I on sheet2
        condition_cell = sheet2['J' + str(late_squad_cell.row)]
        # Get the value of the condition cell
        condition = condition_cell.value
        if condition:
            team_cell = sheet2['G' + str(late_squad_cell.row)]
            # Get the value of the team cell
            team = team_cell.value
            # Check if the team is already in the dictionary
            if late_squad_cell.value in teams:
                # If it is, append the cell to the corresponding list
                teams[late_squad_cell.value].append(team)
            else:
                # If it's not, create a new key in the dictionary with the team as the key and the late squad cell as the value
                teams[late_squad_cell.value] = [team]
        if str(condition) == 'None':
            if late_squad_cell.value in teams:
                # If it is, append the cell to the corresponding list
                teams[late_squad_cell.value].append('None')
            else:
                # If it's not, create a new key in the dictionary with the team as the key and the late squad cell as the value
                teams[late_squad_cell.value] = ['None']
    return teams


    
def compare_teams(matching_cells, late_squad_cells):
    teams_late_dates = {}
    matching_cells_corrected = []
    late_squad_corrected= []
   # Load the workbook

    
    for coordinate in matching_cells:
        #print(coordinate.coordinate)
        if 'D' in coordinate.coordinate:
            sprint_type = 'sprint1'
        if 'E' in coordinate.coordinate:
            sprint = ['lol']
            sprint_type = 'sprint2'
        if 'F' in coordinate.coordinate:
            sprint = ['test']
            sprint_type = 'sprint3'


        matching_cells_corrected.append((coordinate.value , sprint_type))

        for late_squad in late_squad_cells:
            late_squad_corrected.append(late_squad.value)




    teams_late_dates = {key: value for key, value in zip(late_squad_corrected, matching_cells_corrected)}
    return teams_late_dates

def find_cells(teams_late_dates, file_path, tracker_sheet):

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[tracker_sheet]
    date_row = sheet[2]
    team_and_date = {}

    testList = []

    #find the correct date cell on the sheet


    #find the correct team cell in the sheet

    for team in teams_late_dates.keys():
        for cell_c in sheet['C']:
            if cell_c.value:
                testList.append(cell_c.value)

            if cell_c.value == team:

                for cell in date_row:
                        if cell.value == teams_late_dates[team][0]:
                            if cell_c.coordinate in team_and_date:
                                team_and_date[cell_c.coordinate].append(cell.coordinate)
                                team_and_date[cell_c.coordinate].append(teams_late_dates[team][1])

                            else:
                                team_and_date[cell_c.coordinate] = [cell.coordinate]
                                team_and_date[cell_c.coordinate].append(teams_late_dates[team][1])

    return team_and_date




def find_team_role(team_and_date,file_path, team_sheet, tracker_sheet):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(team_sheet)
    tracker_sheet = wb.get_sheet_by_name(tracker_sheet)

    team_roles = {}

    for key in team_and_date.keys():
        key = tracker_sheet[key].value
        for row in range(1, sheet.max_row + 1):

            if key == sheet.cell(row=row, column=4).value:

                if str((sheet.cell(row=row, column=10).value)) != 'None':

                    if key in team_roles:
                        team_roles[key].append((sheet.cell(row=row, column=7).value))
                        
                    else:
                        team_roles[key] = [key]
                        team_roles[key].append((sheet.cell(row=row, column=7).value))
                
                else:
                    pass


    team_roles = {k: v[1:] for k, v in team_roles.items()}

    return team_roles



def colour_team(file_path, tracker_sheet, team_and_date, team_roles, sprint1):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(tracker_sheet)
    default_roles = {}
    

    for key in team_and_date.keys():
        start_row = sheet[key].row
        for row in sheet.iter_rows(min_row=start_row, min_col=7, max_col=7, max_row=start_row+20):
            for cell in row:
                if key in default_roles.keys():
                    if str(cell.value) != 'None' and str(cell.value) :
                        default_roles[key].append(cell.coordinate)
                else:
                    default_roles[key] = [key]
    default_roles = {k: v[1:] for k, v in default_roles.items()}

        
    for team in team_and_date:
        team = sheet[team]
        for team_name in team_roles:
            if team_name == team.value:

                team_and_date[team.coordinate].append(team_roles[team_name])

    return default_roles, team_and_date
            

def colour_now(file_path, tracker_sheet,default_roles, team_date_role):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name(tracker_sheet)
    sprint4 = []


    for squad_cell_to_fill in team_date_role:
        
        if team_date_role[squad_cell_to_fill][1] == 'sprint1':
            team_date_role[squad_cell_to_fill].append(['Designer', 'Senior Dev', 'Dev', 'Test'])

        if team_date_role[squad_cell_to_fill][1] == 'sprint2':
            team_date_role[squad_cell_to_fill].append(['Dev', 'Dev', 'Dev','Test', 'Test', 'Senior Dev', 'Designer', 'Designer' ])

        if team_date_role[squad_cell_to_fill][1] == 'sprint3':
            team_date_role[squad_cell_to_fill].append(['Designer', 'Designer', 'Senior Dev', 'Dev', 'Dev', 'Dev', 'Test', 'Test', 'Scrum Master'])


        if team_date_role[squad_cell_to_fill][1] == 'sprint4':
            team_date_role[squad_cell_to_fill].append(sprint4)



    for squad_cell in team_date_role:
        
        start_row = sheet[squad_cell].row
        end_row = start_row + 20
        for default_row in range(start_row,end_row+1):
            cell_value = sheet.cell(row=default_row,column=7)
            red_fill = PatternFill(start_color="FF0000" , fill_type="solid")
            green_fill = PatternFill(start_color="0000FF00" , fill_type="solid")
            cell_string = str(cell_value.coordinate)
            row = re.findall('\d+', cell_string)
            row =  row[0]
            row2 = str(int(row) +1)
            column = " ".join(re.findall("[a-zA-Z]+", team_date_role[squad_cell][0] )) 
            
            if sheet[column + row].fill.start_color.index == '00000000':

                if str(cell_value.value) in team_date_role[squad_cell][3] and str(cell_value.value) in team_date_role[squad_cell][2]:
                    sheet[column + row].fill = green_fill
                    print(squad_cell)
                    team_date_role[squad_cell][3].remove(str(cell_value.value))
                    team_date_role[squad_cell][2].remove(str(cell_value.value))
                    
                if str(cell_value.value) in team_date_role[squad_cell][3] and str(cell_value.value) in team_date_role[squad_cell][2]:
                    if sheet[column + row].fill.start_color.index == '0000FF00':
                        sheet[column + row2].fill = green_fill
                        team_date_role[squad_cell][2].remove(str(cell_value.value))
                        team_date_role[squad_cell][3].remove(str(cell_value.value))


    for squad_cell in team_date_role:
        start_row = sheet[squad_cell].row
        end_row = start_row + 20
        for default_row in range(start_row,end_row+1):
                
                if sheet['G' + str(default_row)].value in team_date_role[squad_cell][3]:

                    cell = sheet.cell(row=default_row,column=7)
                    cell = str(cell.coordinate)
                    row = re.findall('\d+', cell_string)
                    default_row =  str(default_row)
                    column = " ".join(re.findall("[a-zA-Z]+", team_date_role[squad_cell][0] )) 

                    if sheet[column + default_row].fill.start_color.index == '00000000':
                        sheet[column + default_row].fill = red_fill
                        team_date_role[squad_cell][3].remove(sheet['G' + str(default_row)].value)
                            
        
    wb.save("colored_cell.xlsx")


if __name__ == '__main__':
    file_path = '/home/ali/Desktop/Talktalk/Resourcing and Onboarding Tracker Latest.xlsx'
    dates_sheet_name = 'DevSheet'
    team_sheet = 'ResourceOnboarding Tracker'
    tracker_sheet = "Resources by squad"
    sprint1 = ['Dev', 'Dev', 'Dev','Test', 'Test', 'Senior Dev', 'Designer', 'Designer' ]
    sprint2 = []
    sprint3 = []
    delay_cell_location = []
    sprint =[]

    late_squad_cells, matching_cells = check_dates(file_path, dates_sheet_name)



    if late_squad_cells:
        late_squads = check_team(file_path, team_sheet,late_squad_cells)
        teams_late_dates = compare_teams(matching_cells, late_squad_cells)
        team_and_date = find_cells(teams_late_dates, file_path, tracker_sheet)
        team_roles = find_team_role(team_and_date, file_path, team_sheet, tracker_sheet)
        default_role, team_date_role = colour_team(file_path, tracker_sheet, team_and_date, team_roles, sprint1)
        colour_now(file_path, tracker_sheet,default_role, team_date_role)

